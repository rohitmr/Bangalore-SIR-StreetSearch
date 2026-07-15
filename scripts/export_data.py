#!/usr/bin/env python3
"""Export a constituency's 2002 electoral roll index from xlsx into JSON for the static site.

Usage:
    python3 export_data.py --constituency A087 --xlsx ../Jayamahal_A087_2002_Road_Part_Index_Second_Pass.xlsx
"""
import argparse
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = Path(__file__).resolve().parents[1] / "data"
CONSTITUENCIES_FILE = DATA_DIR / "constituencies.json"


def load_constituency(constituency_id):
    registry = json.loads(CONSTITUENCIES_FILE.read_text())
    for c in registry:
        if c["id"] == constituency_id:
            return c
    raise SystemExit(f"Unknown constituency '{constituency_id}' — check {CONSTITUENCIES_FILE}")


def guess_xlsx_path(constituency):
    name = constituency["name"].replace(" ", "")
    guess = ROOT / f"{name}_{constituency['id']}_2002_Road_Part_Index_Second_Pass.xlsx"
    return guess


def sheet_rows(ws):
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    for r in rows:
        if all(v is None for v in r):
            continue
        yield dict(zip(header, r))


def parse_part_numbers(value):
    if not value:
        return []
    return [int(x.strip()) for x in str(value).split(",") if x.strip()]


def first_present(row, *keys):
    """Different constituencies' workbooks have renamed some columns — try each in turn."""
    for k in keys:
        v = row.get(k)
        if v is not None:
            return v
    return None


def count_by_part_number(ws):
    """Tally rows per Part Number in a detail sheet (e.g. Accepted Entries, Needs Review) —
    used as a fallback when Part Summary doesn't have a pre-tallied count column."""
    counts = {}
    for row in sheet_rows(ws):
        pn = row.get("Part Number")
        if pn is None:
            continue
        counts[str(pn)] = counts.get(str(pn), 0) + 1
    return counts


def pin_status_by_road(ws):
    """Some workbooks' Accepted Entries sheet records a Map Pin Status per road
    (e.g. "Suppressed — search link only") — the verification pass's own call on
    whether a road's location is confident enough to plot, separate from whether
    the road/part match itself is confident. Geocoding must respect this instead
    of asserting a pin Nominatim happens to return for a name it was never asked
    to actually verify the location of."""
    statuses = {}
    for row in sheet_rows(ws):
        road = row.get("Road / Locality")
        status = row.get("Map Pin Status")
        if road is None or status is None:
            continue
        statuses[road] = status
    return statuses


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--constituency", required=True, help="Constituency ID, e.g. A087")
    parser.add_argument("--xlsx", help="Path to the constituency's xlsx workbook (default: guessed from its name)")
    args = parser.parse_args()

    constituency = load_constituency(args.constituency)
    xlsx_path = Path(args.xlsx) if args.xlsx else guess_xlsx_path(constituency)
    if not xlsx_path.exists():
        raise SystemExit(f"xlsx not found at {xlsx_path} — pass --xlsx explicitly")

    official_pdf_template = (
        f"https://bbmp-sir-2002.s3.us-east-1.amazonaws.com/{constituency['id']}-PDFs/{{code}}.pdf"
    )
    out_dir = DATA_DIR / constituency["id"]

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    accepted_counts = count_by_part_number(wb["Accepted Entries"]) if "Accepted Entries" in wb.sheetnames else {}
    needs_review_counts = count_by_part_number(wb["Needs Review"]) if "Needs Review" in wb.sheetnames else {}
    pin_statuses = pin_status_by_road(wb["Accepted Entries"]) if "Accepted Entries" in wb.sheetnames else {}

    road_index = []
    for row in sheet_rows(wb["Road Index"]):
        parts = parse_part_numbers(row.get("Part Number(s)"))
        road_index.append({
            "road": row.get("Road / Locality"),
            "parts": parts,
            "partCount": row.get("Part Count"),
            "matchedEntries": row.get("Matched Entries"),
            "minConfidence": row.get("Minimum Confidence"),
            "status": row.get("Status"),
            "numberedRoadDetails": first_present(row, "Numbered Roads / Details", "Numbered Road / Coverage Details"),
            "mapAreaCheck": row.get("Map / Area Check"),
            "clusterZone": row.get("Cluster Zone(s)"),
            "evidenceUrl": row.get("Evidence URL"),
            "googleMapsSearch": row.get("Google Maps Search"),
            "pinStatus": pin_statuses.get(row.get("Road / Locality")),
        })

    part_summary = []
    for row in sheet_rows(wb["Part Summary"]):
        part_no = row.get("Part Number")
        code = row.get("Part Code")
        accepted = row.get("Accepted Entries")
        if accepted is None:
            accepted = accepted_counts.get(str(part_no), 0)
        needs_review_count = row.get("Needs Review")
        if needs_review_count is None:
            needs_review_count = needs_review_counts.get(str(part_no), 0)
        part_summary.append({
            "partNumber": part_no,
            "partCode": code,
            "roads": row.get("Roads / Localities"),
            "numberedRoadDetails": first_present(row, "Numbered Road Details", "Numbered Road / Coverage Details"),
            "acceptedEntries": accepted,
            "historicalOnlyEntries": row.get("Historical-only Entries") or 0,
            "needsReview": needs_review_count,
            "excludedOcr": row.get("Excluded OCR") or 0,
            "clusterZone": row.get("Cluster Zone"),
            "sourceImage": row.get("Source Image"),
            "officialPdfUrl": official_pdf_template.format(code=code),
        })

    # BBMP Gazetteer columns vary by constituency's workbook — pass through as-is
    # rather than forcing a fixed shape (this file isn't consumed by the site,
    # only kept for the user's own cross-referencing).
    bbmp_gazetteer = list(sheet_rows(wb["BBMP Gazetteer"]))

    needs_review = list(sheet_rows(wb["Needs Review"]))

    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "road_index.json").write_text(json.dumps(road_index, ensure_ascii=False, indent=2))
    (out_dir / "part_summary.json").write_text(json.dumps(part_summary, ensure_ascii=False, indent=2))
    (out_dir / "bbmp_gazetteer.json").write_text(json.dumps(bbmp_gazetteer, ensure_ascii=False, indent=2, default=str))
    (out_dir / "needs_review.json").write_text(json.dumps(needs_review, ensure_ascii=False, indent=2, default=str))

    print(f"{constituency['id']} ({constituency['name']}) -> {out_dir}")
    print(f"road_index: {len(road_index)} entries")
    print(f"part_summary: {len(part_summary)} entries")
    print(f"bbmp_gazetteer: {len(bbmp_gazetteer)} entries")
    print(f"needs_review: {len(needs_review)} entries")


if __name__ == "__main__":
    main()
